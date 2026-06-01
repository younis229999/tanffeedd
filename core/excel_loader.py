"""
قراءة وكتابة ملفات Excel.

مبادئ مهمة:
- لا نكتب فوق الملف الأصلي إطلاقاً — ننشئ ملفات جديدة فقط.
- نقرأ الأعمدة حسب حروفها (A/E/H/I) القابلة للتعديل من الإعدادات.
- ملف الإخراج باتجاه LTR (النمط الإنجليزي): ``sheet_view.rightToLeft = False``.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# الترتيب القياسي لأعمدة الإخراج وعناوينها (إنجليزية، LTR).
OUTPUT_COLUMNS = ["folder", "amount", "name", "iban"]
OUTPUT_HEADERS = {
    "folder": "Folder No.",
    "amount": "Amount",
    "name": "Name",
    "iban": "IBAN",
}


@dataclass
class LoadedData:
    """البيانات المقروءة من ملف الإدخال بعد ربط الأعمدة."""

    folders: List[object] = field(default_factory=list)
    amounts: List[object] = field(default_factory=list)
    names: List[str] = field(default_factory=list)
    ibans: List[str] = field(default_factory=list)
    source_path: str = ""

    def __len__(self) -> int:
        return len(self.folders)

    def as_rows(self) -> List[Dict[str, object]]:
        """تحويل البيانات إلى قائمة صفوف (قواميس) للعرض/المعالجة."""
        rows = []
        for i in range(len(self)):
            rows.append(
                {
                    "folder": self.folders[i],
                    "amount": self.amounts[i],
                    "name": self.names[i],
                    "iban": self.ibans[i],
                }
            )
        return rows


class ExcelLoadError(Exception):
    """خطأ في قراءة ملف Excel (تالف أو أعمدة غير متطابقة)."""


def _col_to_index(letter: str) -> int:
    """تحويل حرف العمود (A, E, ...) إلى فهرس 0-based."""
    return column_index_from_string(str(letter).strip().upper()) - 1


def _clean_cell_text(value: object) -> str:
    """تحويل قيمة خلية إلى نص نظيف (مع معالجة NaN والفراغات)."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def load_excel(path: str, columns: Dict[str, str], has_header: bool = True) -> LoadedData:
    """
    قراءة ملف Excel وربط الأعمدة المطلوبة.

    المعطيات:
      path: مسار ملف الإدخال (لا يُعدّل).
      columns: قاموس {folder, amount, name, iban} -> حرف العمود.
      has_header: هل الصف الأول عنوان (يُتخطى)؟

    تعيد LoadedData. ترفع ExcelLoadError عند الفشل.
    """
    p = Path(path)
    if not p.exists():
        raise ExcelLoadError(f"الملف غير موجود: {path}")

    try:
        # نقرأ كل الأعمدة كنص خام دون افتراض عناوين، ثم نختار بالحرف.
        raw = pd.read_excel(p, header=None, dtype=object, engine="openpyxl")
    except Exception as exc:  # noqa: BLE001 — نريد رسالة عربية موحّدة.
        raise ExcelLoadError(
            f"تعذّر فتح الملف. تأكد أنه ملف Excel سليم (.xlsx).\nالتفاصيل: {exc}"
        ) from exc

    if raw.empty:
        raise ExcelLoadError("الملف فارغ — لا توجد بيانات للمعالجة.")

    # تخطّي صف العنوان عند الحاجة.
    if has_header and len(raw) > 0:
        raw = raw.iloc[1:].reset_index(drop=True)

    if raw.empty:
        raise ExcelLoadError("لا توجد صفوف بيانات بعد صف العنوان.")

    try:
        idx = {key: _col_to_index(col) for key, col in columns.items()}
    except (ValueError, KeyError) as exc:
        raise ExcelLoadError(f"إعداد الأعمدة غير صحيح: {exc}") from exc

    max_needed = max(idx.values())
    if raw.shape[1] <= max_needed:
        raise ExcelLoadError(
            "عدد الأعمدة في الملف أقل من المتوقع. "
            f"المطلوب على الأقل العمود {get_column_letter(max_needed + 1)}، "
            f"بينما الملف يحوي {raw.shape[1]} عمود."
        )

    data = LoadedData(source_path=str(p))
    for _, row in raw.iterrows():
        folder = row.iloc[idx["folder"]]
        amount = row.iloc[idx["amount"]]
        name = _clean_cell_text(row.iloc[idx["name"]])
        iban = _clean_cell_text(row.iloc[idx["iban"]]).replace(" ", "").upper()

        # تخطّي الصفوف الفارغة تماماً.
        if not _clean_cell_text(folder) and not name and not iban and \
                _clean_cell_text(amount) == "":
            continue

        data.folders.append(_clean_cell_text(folder))
        data.amounts.append(amount)
        data.names.append(name)
        data.ibans.append(iban)

    if len(data) == 0:
        raise ExcelLoadError("لم يُعثر على صفوف بيانات صالحة في الملف.")

    return data


def parse_amount(value: object) -> Optional[float]:
    """
    تحويل قيمة مبلغ إلى رقم.

    يتعامل مع الفواصل (1,000,000) والمسافات والنص. يعيد None إذا تعذّر.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and pd.isna(value):
            return None
        return float(value)
    # تطبيع الأرقام العربية-الهندية/الفارسية إلى غربية قبل التحويل.
    digit_map = {ord(c): str(i % 10) for i, c in enumerate("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹")}
    text = str(value).translate(digit_map).strip()
    # إزالة الفواصل والمسافات والفاصلة العربية.
    text = text.replace(",", "").replace("،", "").replace(" ", "").replace("٬", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def write_output_excel(
    rows: List[Dict[str, object]],
    out_path: str,
    directorate_name: str = "",
) -> str:
    """
    كتابة ملف Excel النهائي المنظّف باتجاه LTR.

    المعطيات:
      rows: صفوف الناتج (قواميس فيها folder, amount, name, iban).
      out_path: مسار الحفظ.
      directorate_name: اسم المديرية (لا يُكتب في الورقة، للتوافق فقط).

    تعيد مسار الملف المكتوب.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Localization"

    # اتجاه الورقة LTR (النمط الإنجليزي) — مطلب صريح لملف الإكسل.
    ws.sheet_view.rightToLeft = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    # صف العناوين.
    for col_i, key in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_i, value=OUTPUT_HEADERS[key])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # الصفوف.
    for r_i, row in enumerate(rows, start=2):
        ws.cell(row=r_i, column=1, value=row.get("folder", ""))
        amount = row.get("amount", 0)
        amt_cell = ws.cell(row=r_i, column=2, value=amount)
        amt_cell.number_format = "0"   # بدون فواصل آلاف.
        ws.cell(row=r_i, column=3, value=row.get("name", ""))
        ws.cell(row=r_i, column=4, value=row.get("iban", ""))

    # ضبط عرض الأعمدة.
    widths = {1: 16, 2: 16, 3: 32, 4: 30}
    for col_i, width in widths.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_output_filename(directory: str, prefix: str = "altanfith") -> str:
    """توليد اسم ملف ناتج يتضمن التاريخ والوقت."""
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return str(Path(directory) / f"{prefix}_{stamp}.xlsx")
